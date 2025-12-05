"""
FastAPI应用主入口文件 - 兼容版本
保持与原有main.py的兼容性，同时使用新的模块化结构
"""
import os
import sys
from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware

# 添加src目录到Python路径
BASE_DIR = os.path.dirname(__file__)
src_path = os.path.join(BASE_DIR, "src")
if src_path not in sys.path:
    sys.path.append(src_path)

# 导入新的模块化组件
try:
    from src.main import app as new_app
    # 如果新的模块化应用可用，使用它
    app = new_app
except ImportError:
    # 如果新模块不可用，回退到原有实现
    from src.config.config import config
    from src.dao.supabase_client import get_client
    
    # 创建FastAPI应用
    app = FastAPI(
        title="Price Memory API",
        description="价格记忆 - 商品价格监控与分析API",
        version="1.0.0"
    )
    
    # 添加CORS中间件
    app.add_middleware(
        CORSMiddleware,
        allow_origins=["*"],
        allow_credentials=True,
        allow_methods=["*"],
        allow_headers=["*"],
    )
    
    # 初始化数据库连接
    try:
        SB = get_client()
    except Exception:
        SB = None

def get_auth_uid(user_id: int) -> Optional[str]:
    if not SB:
        return None
    res = SB.table("users").select("auth_uid").eq("id", user_id).limit(1).execute()
    rows = getattr(res, "data", None) or []
    uid = (rows[0] or {}).get("auth_uid") if rows else None
    return uid

import sqlite3
def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("CREATE TABLE IF NOT EXISTS products (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL, url TEXT NOT NULL, category TEXT, last_updated TEXT)")
    cur.execute("CREATE TABLE IF NOT EXISTS prices (id INTEGER PRIMARY KEY AUTOINCREMENT, product_id INTEGER NOT NULL, price REAL NOT NULL, created_at TEXT NOT NULL)")
    cur.execute("CREATE TABLE IF NOT EXISTS tasks (id INTEGER PRIMARY KEY AUTOINCREMENT, product_id INTEGER, status TEXT NOT NULL, priority INTEGER DEFAULT 0, created_at TEXT NOT NULL, updated_at TEXT NOT NULL, started_at TEXT, completed_at TEXT)")
    cur.execute("CREATE TABLE IF NOT EXISTS users (id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT UNIQUE, display_name TEXT, email TEXT, created_at TEXT, api_key TEXT, plan TEXT, quota_exports_per_day INTEGER, exports_used_today INTEGER, last_quota_reset TEXT, quota_tasks_per_day INTEGER, tasks_created_today INTEGER, last_tasks_quota_reset TEXT)")
    cur.execute("CREATE TABLE IF NOT EXISTS pools (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT, is_public INTEGER DEFAULT 1)")
    cur.execute("CREATE TABLE IF NOT EXISTS pool_products (id INTEGER PRIMARY KEY AUTOINCREMENT, pool_id INTEGER NOT NULL, product_id INTEGER NOT NULL)")
    conn.commit()
    conn.close()
try:
    src_path = os.path.join(BASE_DIR, "src")
    if src_path not in sys.path:
        sys.path.append(src_path)
    from src.dao.supabase_client import get_client
    SB = get_client()
except Exception:
    SB = None

def get_conn():
    import sqlite3
    conn = sqlite3.connect(DB_PATH)
    try:
        conn.row_factory = sqlite3.Row
    except Exception:
        pass
    return conn

def init_db():
    import sqlite3
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("CREATE TABLE IF NOT EXISTS products (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT, url TEXT, category TEXT, last_updated TEXT)")
    cur.execute("CREATE TABLE IF NOT EXISTS prices (id INTEGER PRIMARY KEY AUTOINCREMENT, product_id INTEGER, price REAL, created_at TEXT)")
    cur.execute("CREATE TABLE IF NOT EXISTS tasks (id INTEGER PRIMARY KEY AUTOINCREMENT, product_id INTEGER, status TEXT, priority INTEGER, created_at TEXT, updated_at TEXT, scheduled_at TEXT, started_at TEXT, completed_at TEXT, created_by_user_id INTEGER)")
    conn.commit()
    conn.close()

def now_iso() -> str:
    return datetime.datetime.utcnow().replace(microsecond=0).isoformat() + "Z"

def ok(data: Any, message: str = "操作成功"):
    return {"success": True, "data": data, "message": message, "timestamp": now_iso()}

def error_response(status_code: int, code: str, message: str, details: Optional[List[Any]] = None):
    if os.getenv("STRICT_HTTP") == "1":
        return JSONResponse(status_code=status_code, content={"success": False, "error": {"code": code, "message": message, "details": details or []}, "timestamp": now_iso()})
    return {"success": False, "error": {"code": code, "message": message, "details": details or []}, "timestamp": now_iso()}


def get_user_by_api_key(api_key: Optional[str]) -> Optional[dict]:
    if not api_key or not isinstance(api_key, str):
        return None
    if SB:
        res = SB.table("users").select("*").eq("api_key", api_key).limit(1).execute()
        data = getattr(res, "data", None) or []
        return data[0] if data else None
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id, api_key, quota_exports_per_day, exports_used_today, last_quota_reset FROM users WHERE api_key = ? LIMIT 1", (api_key,))
    r = cur.fetchone()
    conn.close()
    return {"id": r[0], "api_key": r[1], "quota_exports_per_day": r[2], "exports_used_today": r[3], "last_quota_reset": r[4]} if r else None

def reset_user_quota_if_needed(user_id: int):
    today = datetime.datetime.utcnow().date().isoformat()
    res = SB.table("users").select("exports_used_today,last_quota_reset").eq("id", user_id).limit(1).execute()
    data = getattr(res, "data", None) or []
    if not data:
        return
    last = data[0].get("last_quota_reset")
    if last != today:
        SB.table("users").update({"exports_used_today": 0, "last_quota_reset": today}).eq("id", user_id).execute()

def row_to_product(r: dict) -> dict:
    return {"id": r.get("id"), "name": r.get("name"), "url": r.get("url"), "category": r.get("category"), "last_updated": r.get("updated_at")}

def row_to_price(r: dict) -> dict:
    return {"id": r.get("id"), "product_id": r.get("product_id"), "price": r.get("price"), "created_at": r.get("created_at")}

def create_product(name: str, url: str, category: Optional[str] = None) -> int:
    if SB:
        try:
            now = now_iso()
            res = SB.table("products").insert({"name": name, "url": url, "category": category, "updated_at": now}).select("id").execute()
            data = getattr(res, "data", None) or []
            if data:
                return int(data[0]["id"]) if data else 0
        except Exception:
            pass
    conn = get_conn()
    cur = conn.cursor()
    now = now_iso()
    cur.execute("INSERT INTO products(name, url, category, last_updated) VALUES(?, ?, ?, ?)", (name, url, category, now))
    conn.commit()
    pid = cur.lastrowid
    conn.close()
    return int(pid)

def get_product(product_id: int) -> Optional[dict]:
    if SB:
        try:
            res = SB.table("products").select("*").eq("id", product_id).limit(1).execute()
            data = getattr(res, "data", None) or []
            if data:
                r = data[0]
                return {"id": r.get("id"), "name": r.get("name"), "url": r.get("url"), "category": r.get("category"), "last_updated": r.get("updated_at")}
        except Exception:
            pass
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM products WHERE id = ?", (product_id,))
    r = cur.fetchone()
    conn.close()
    if not r:
        return None
    return {"id": r["id"], "name": r["name"], "url": r["url"], "category": r["category"], "last_updated": r["last_updated"]}

class ProductCreate(BaseModel):
    name: str
    url: str
    category: Optional[str] = None

class ListingRequest(BaseModel):
    url: str
    max_items: int = 50

class TaskCreate(BaseModel):
    product_id: Optional[int] = None
    priority: Optional[int] = 0

class UserCreate(BaseModel):
    username: str
    display_name: Optional[str] = None
    email: Optional[str] = None

class FollowCreate(BaseModel):
    product_id: int

class PushCreate(BaseModel):
    recipient_id: int
    product_id: int
    message: Optional[str] = None

class PushUpdate(BaseModel):
    status: str

class PoolAddProduct(BaseModel):
    product_id: int

class SelectFromPoolBody(BaseModel):
    product_id: int

class CollectionCreate(BaseModel):
    name: str
    owner_user_id: int

class CollectionAddProduct(BaseModel):
    product_id: int

class CollectionShare(BaseModel):
    user_id: int
    role: Optional[str] = "editor"

class AlertCreate(BaseModel):
    user_id: int
    product_id: int
    rule_type: str
    threshold: Optional[float] = None
    percent: Optional[float] = None
    channel: Optional[str] = None
    cooldown_minutes: Optional[int] = None
    target: Optional[str] = None

class AlertStatusUpdate(BaseModel):
    status: str

class PreferencesUpdate(BaseModel):
    trend_ma_window: Optional[int] = None
    trend_bb_on: Optional[bool] = None

app = FastAPI()
try:
    from fastapi.middleware.cors import CORSMiddleware
    app.add_middleware(
        CORSMiddleware,
        allow_origins=["*"],
        allow_credentials=True,
        allow_methods=["*"],
        allow_headers=["*"],
    )
except Exception:
    pass
router = APIRouter(prefix="/api/v1")
RATE_LIMIT: Dict[str, float] = {}
def _is_node_paused() -> bool:
    try:
        if str(os.environ.get("NODE_PAUSED") or "").strip() == "1":
            return True
        import socket
        name = os.environ.get("NODE_NAME") or f"node-{socket.gethostname()}"
        src_path = os.path.join(BASE_DIR, "src")
        if src_path not in sys.path:
            sys.path.append(src_path)
        from src.dao.supabase_client import get_client
        client = get_client()
        if not client:
            return False
        res = client.table("runtime_nodes").select("status").eq("name", name).limit(1).execute()
        rows = getattr(res, "data", None) or []
        if rows:
            return str(rows[0].get("status")) == "paused"
        return False
    except Exception:
        return False

@app.on_event("startup")
def on_startup():
    sys.path.append(os.path.join(BASE_DIR, "src"))
    try:
        from runtime.node_runtime import NodeRuntime
        rt = NodeRuntime()
        rt.start()
    except Exception:
        pass

@router.get("/system/status")
def system_status():
    if SB:
        total = getattr(SB.table("tasks").select("id", count="exact").execute(), "count", 0) or 0
        completed = getattr(SB.table("tasks").select("id", count="exact").eq("status", "completed").execute(), "count", 0) or 0
        pending = getattr(SB.table("tasks").select("id", count="exact").eq("status", "pending").execute(), "count", 0) or 0
        today = datetime.datetime.utcnow().date().isoformat()
        today_count = getattr(SB.table("tasks").select("id", count="exact").gte("created_at", today).execute(), "count", 0) or 0
        return ok({"health": "ok", "today_tasks": today_count, "total_tasks": total, "completed_tasks": completed, "pending_tasks": pending})
    conn = get_conn()
    cur = conn.cursor()
    total = cur.execute("SELECT COUNT(*) FROM tasks").fetchone()[0]
    completed = cur.execute("SELECT COUNT(*) FROM tasks WHERE status = 'completed'").fetchone()[0]
    pending = cur.execute("SELECT COUNT(*) FROM tasks WHERE status = 'pending'").fetchone()[0]
    today = datetime.datetime.utcnow().date().isoformat()
    today_count = cur.execute("SELECT COUNT(*) FROM tasks WHERE substr(created_at,1,10) = ?", (today,)).fetchone()[0]
    conn.close()
    return ok({"health": "ok", "today_tasks": today_count, "total_tasks": total, "completed_tasks": completed, "pending_tasks": pending})

@router.get("/auth/permissions")
def auth_permissions(api_key: Optional[str] = Header(None)):
    perms: List[Dict[str, str]] = []
    perms.append({"resource": "products", "action": "export"})
    perms.append({"resource": "collections", "action": "share"})
    perms.append({"resource": "collections", "action": "export"})
    perms.append({"resource": "public-pool", "action": "select"})
    perms.append({"resource": "pushes", "action": "update"})
    return ok(perms)

@router.get("/products")
def list_products(page: int = Query(1, ge=1), size: int = Query(20, ge=1, le=100)):
    offset = (page - 1) * size
    sel = SB.table("products").select("*", count="exact").order("id", desc=True).range(offset, offset + size - 1).execute()
    items = [{"id": r.get("id"), "name": r.get("name"), "url": r.get("url"), "category": r.get("category"), "last_updated": r.get("updated_at")} for r in (getattr(sel, "data", None) or [])]
    total = getattr(sel, "count", 0) or len(items)
    return ok({"items": items, "page": page, "size": size, "total": total, "pages": math.ceil(total / size) if size else 0})

@router.get("/products/search")
def search_products(page: int = Query(1, ge=1), size: int = Query(20, ge=1, le=100), search: Optional[str] = None, category: Optional[str] = None, sort_by: Optional[str] = None, sort_order: Optional[str] = None):
    q = SB.table("products").select("*", count="exact")
    if search:
        q = q.ilike("name", f"%{search}%")
    if category:
        q = q.eq("category", category)
    if sort_by in {"name", "updated_at"}:
        q = q.order(sort_by, desc=(sort_order == "desc"))
    offset = (page - 1) * size
    res = q.range(offset, offset + size - 1).execute()
    items = [{"id": r.get("id"), "name": r.get("name"), "url": r.get("url"), "category": r.get("category"), "last_updated": r.get("updated_at")} for r in (getattr(res, "data", None) or [])]
    total = getattr(res, "count", 0) or len(items)
    return ok({"items": items, "page": page, "size": size, "total": total, "pages": math.ceil(total / size) if size else 0})

@router.post("/products")
def create_product_endpoint(body: ProductCreate):
    try:
        pid = create_product(body.name, body.url, body.category)
        return ok({"id": pid, "name": body.name, "url": body.url, "category": body.category})
    except Exception as e:
        return error_response(500, "INTERNAL_ERROR", str(e))

@router.patch("/products/{product_id}")
def update_product_endpoint(product_id: int, body: ProductCreate):
    conn = get_conn()
    cur = conn.cursor()
    now = now_iso()
    cur.execute("UPDATE products SET name = ?, url = ?, category = ?, last_updated = ? WHERE id = ?", (body.name, body.url, body.category, now, product_id))
    conn.commit()
    cur.execute("SELECT * FROM products WHERE id = ?", (product_id,))
    r = cur.fetchone()
    conn.close()
    if not r:
        return error_response(404, "NOT_FOUND", "资源不存在")
    return ok(row_to_product(r))

@router.delete("/products/{product_id}")
def delete_product_endpoint(product_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM prices WHERE product_id = ?", (product_id,))
    cur.execute("DELETE FROM tasks WHERE product_id = ?", (product_id,))
    cur.execute("DELETE FROM products WHERE id = ?", (product_id,))
    conn.commit()
    conn.close()
    return ok({"id": product_id})

@router.get("/products/{product_id}")
def product_detail(product_id: int):
    p = get_product(product_id)
    if not p:
        return error_response(404, "NOT_FOUND", "资源不存在")
    stats = SB.rpc("rpc_product_stats", {"product_id": product_id}).execute()
    rows = getattr(stats, "data", None) or []
    if rows:
        r = rows[0]
        p["stats"] = {"count": r.get("count"), "max_price": r.get("max_price"), "min_price": r.get("min_price"), "avg_price": r.get("avg_price")}
    else:
        p["stats"] = {"count": 0, "max_price": None, "min_price": None, "avg_price": None}
    return ok(p)

@router.get("/products/{product_id}/prices")
def product_prices(product_id: int, start_date: Optional[str] = None, end_date: Optional[str] = None):
    p = get_product(product_id)
    if not p:
        return error_response(404, "NOT_FOUND", "资源不存在")
    q = SB.table("prices").select("*").eq("product_id", product_id).order("created_at", desc=True)
    if start_date:
        q = q.gte("created_at", start_date)
    if end_date:
        q = q.lte("created_at", end_date + " 23:59:59")
    res = q.execute()
    items = [{"id": r.get("id"), "product_id": r.get("product_id"), "price": r.get("price"), "created_at": r.get("created_at")} for r in (getattr(res, "data", None) or [])]
    return ok(items)

@router.get("/products/{product_id}/trend")
def product_trend(product_id: int, start_date: Optional[str] = None, end_date: Optional[str] = None, granularity: Optional[str] = "daily"):
    p = get_product(product_id)
    if not p:
        return error_response(404, "NOT_FOUND", "资源不存在")
    if granularity not in {"daily", "hourly"}:
        return error_response(400, "VALIDATION_ERROR", "不支持的粒度")
    if granularity == "daily":
        sd = start_date or datetime.datetime.utcnow().date().isoformat()
        ed = end_date or sd
        res = SB.rpc("rpc_product_daily_ohlc", {"product_id": product_id, "start_date": sd, "end_date": ed}).execute()
        rows = getattr(res, "data", None) or []
        series = [{"date": r.get("day"), "open": r.get("open"), "close": r.get("close"), "low": r.get("low"), "high": r.get("high"), "avg": r.get("avg"), "count": r.get("count")} for r in rows]
        return ok({"granularity": granularity, "series": series})
    else:
        sd = start_date or datetime.datetime.utcnow().date().isoformat()
        ed = end_date or sd
        res = SB.rpc("rpc_product_hourly_ohlc", {"product_id": product_id, "start_ts": sd + " 00:00:00", "end_ts": ed + " 23:59:59"}).execute()
        rows = getattr(res, "data", None) or []
        series = [{"date": r.get("hour"), "open": r.get("open"), "close": r.get("close"), "low": r.get("low"), "high": r.get("high"), "avg": r.get("avg"), "count": r.get("count")} for r in rows]
        return ok({"granularity": granularity, "series": series})

@router.get("/products/{product_id}/export")
def export_product_prices(product_id: int, start_date: Optional[str] = None, end_date: Optional[str] = None, api_key: Optional[str] = None):
    p = get_product(product_id)
    if not p:
        return error_response(404, "NOT_FOUND", "资源不存在")
    user = get_user_by_api_key(api_key)
    if user:
        reset_user_quota_if_needed(int(user["id"]))
        quota = user.get("quota_exports_per_day") or 0
        used = user.get("exports_used_today") or 0
        if quota and used >= quota:
            return error_response(429, "QUOTA_EXCEEDED", "导出额度已用尽")
    q = SB.table("prices").select("id,price,created_at").eq("product_id", product_id).order("created_at", desc=True)
    if start_date:
        q = q.gte("created_at", start_date)
    if end_date:
        q = q.lte("created_at", end_date + " 23:59:59")
    rows = getattr(q.execute(), "data", None) or []
    import csv
    import io
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["product_id", "product_name", "url", "category", "price_id", "price", "created_at"])
    for r in rows:
        writer.writerow([product_id, p["name"], p["url"], p["category"], r.get("id"), r.get("price"), r.get("created_at")])
    csv_content = output.getvalue()
    filename = f"product_{product_id}_prices.csv"
    if user:
        SB.table("users").update({"exports_used_today": (used or 0) + 1, "last_quota_reset": datetime.datetime.utcnow().date().isoformat()}).eq("id", int(user["id"])).execute()
    return Response(content=csv_content, media_type="text/csv", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

@router.get("/export")
def export_products(product_ids: str, start_date: Optional[str] = None, end_date: Optional[str] = None, api_key: Optional[str] = Header(None)):
    ids: List[int] = []
    for s in (product_ids or "").split(","):
        s = s.strip()
        if not s:
            continue
        try:
            ids.append(int(s))
        except Exception:
            continue
    if not ids:
        return error_response(400, "VALIDATION_ERROR", "缺少产品ID")
    user = get_user_by_api_key(api_key)
    if user:
        reset_user_quota_if_needed(int(user["id"]))
        quota = user.get("quota_exports_per_day") or 0
        used = user.get("exports_used_today") or 0
        if quota and (used + len(ids)) > quota:
            return error_response(429, "QUOTA_EXCEEDED", "导出额度已用尽")
    import csv
    import io
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["product_id", "product_name", "url", "category", "price_id", "price", "created_at"])
    exported_count = 0
    for pid in ids:
        p = get_product(pid)
        if not p:
            continue
        q = SB.table("prices").select("id,price,created_at").eq("product_id", pid).order("created_at", desc=True)
        if start_date:
            q = q.gte("created_at", start_date)
        if end_date:
            q = q.lte("created_at", end_date + " 23:59:59")
        rows = getattr(q.execute(), "data", None) or []
        for r in rows:
            writer.writerow([pid, p["name"], p["url"], p["category"], r.get("id"), r.get("price"), r.get("created_at")])
        exported_count += 1
    csv_content = output.getvalue()
    if user and exported_count:
        SB.table("users").update({"exports_used_today": (user.get("exports_used_today") or 0) + exported_count, "last_quota_reset": datetime.datetime.utcnow().date().isoformat()}).eq("id", int(user["id"])).execute()
    return Response(content=csv_content, media_type="text/csv", headers={"Content-Disposition": 'attachment; filename="products_export.csv"'})

@router.get("/export/zip")
def export_products_zip(product_ids: str, start_date: Optional[str] = None, end_date: Optional[str] = None, api_key: Optional[str] = Header(None)):
    ids: List[int] = []
    for s in (product_ids or "").split(","):
        s = s.strip()
        if not s:
            continue
        try:
            ids.append(int(s))
        except Exception:
            continue
    if not ids:
        return error_response(400, "VALIDATION_ERROR", "缺少产品ID")
    user = get_user_by_api_key(api_key)
    if user:
        reset_user_quota_if_needed(int(user["id"]))
        quota = user.get("quota_exports_per_day") or 0
        used = user.get("exports_used_today") or 0
        if quota and (used + len(ids)) > quota:
            return error_response(429, "QUOTA_EXCEEDED", "导出额度已用尽")
    import io
    import csv
    import zipfile
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for pid in ids:
            p = get_product(pid)
            if not p:
                continue
            q = SB.table("prices").select("id,price,created_at").eq("product_id", pid).order("created_at", desc=True)
            if start_date:
                q = q.gte("created_at", start_date)
            if end_date:
                q = q.lte("created_at", end_date + " 23:59:59")
            rows = getattr(q.execute(), "data", None) or []
            csv_io = io.StringIO()
            writer = csv.writer(csv_io)
            writer.writerow(["product_id", "product_name", "url", "category", "price_id", "price", "created_at"])
            for r in rows:
                writer.writerow([pid, p["name"], p["url"], p["category"], r.get("id"), r.get("price"), r.get("created_at")])
            zf.writestr(f"product_{pid}_prices.csv", csv_io.getvalue())
    if user:
        SB.table("users").update({"exports_used_today": (user.get("exports_used_today") or 0) + len(ids), "last_quota_reset": datetime.datetime.utcnow().date().isoformat()}).eq("id", int(user["id"])).execute()
    zip_bytes = zip_buffer.getvalue()
    return Response(content=zip_bytes, media_type="application/zip", headers={"Content-Disposition": 'attachment; filename="products_export.zip"'})

@router.get("/export/xlsx")
def export_products_xlsx(product_ids: str, start_date: Optional[str] = None, end_date: Optional[str] = None, api_key: Optional[str] = Header(None)):
    ids: List[int] = []
    for s in (product_ids or "").split(","):
        s = s.strip()
        if not s:
            continue
        try:
            ids.append(int(s))
        except Exception:
            continue
    if not ids:
        return error_response(400, "VALIDATION_ERROR", "缺少产品ID")
    user = get_user_by_api_key(api_key)
    if user:
        reset_user_quota_if_needed(int(user["id"]))
        quota = user.get("quota_exports_per_day") or 0
        used = user.get("exports_used_today") or 0
        if quota and (used + len(ids)) > quota:
            return error_response(429, "QUOTA_EXCEEDED", "导出额度已用尽")
    from openpyxl import Workbook
    import io
    wb = Workbook()
    ws_default = wb.active
    ws_default.title = "Summary"
    ws_default.append(["product_id", "product_name", "url", "category", "price_id", "price", "created_at"])
    for pid in ids:
        p = get_product(pid)
        if not p:
            continue
        title = str(pid)
        try:
            title = (p.get("name") or str(pid))[:31].replace("/", "-")
        except Exception:
            title = str(pid)
        ws = wb.create_sheet(title=title)
        ws.append(["product_id", "product_name", "url", "category", "price_id", "price", "created_at"])
        q = SB.table("prices").select("id,price,created_at").eq("product_id", pid).order("created_at", desc=True)
        if start_date:
            q = q.gte("created_at", start_date)
        if end_date:
            q = q.lte("created_at", end_date + " 23:59:59")
        rows = getattr(q.execute(), "data", None) or []
        for r in rows:
            record = [pid, p["name"], p["url"], p["category"], r.get("id"), r.get("price"), r.get("created_at")]
            ws.append(record)
            ws_default.append(record)
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    if user:
        SB.table("users").update({"exports_used_today": (user.get("exports_used_today") or 0) + len(ids), "last_quota_reset": datetime.datetime.utcnow().date().isoformat()}).eq("id", int(user["id"])).execute()
    return Response(content=content, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": 'attachment; filename="products_export.xlsx"'})

@router.post("/users")
def create_user(body: UserCreate):
    if SB:
        now = now_iso()
        api_key = secrets.token_hex(16)
        plan = "basic"
        quota_exports = 5
        today = datetime.datetime.utcnow().date().isoformat()
        res = SB.table("users").insert({"username": body.username, "display_name": body.display_name, "created_at": now, "email": body.email, "api_key": api_key, "plan": plan, "quota_exports_per_day": quota_exports, "exports_used_today": 0, "last_quota_reset": today}).select("id,username,display_name,created_at,email,api_key,plan,quota_exports_per_day").execute()
        data = getattr(res, "data", None) or []
        if not data:
            return error_response(400, "VALIDATION_ERROR", "用户名已存在")
        r = data[0]
        return ok({"id": r.get("id"), "username": r.get("username"), "display_name": r.get("display_name"), "created_at": r.get("created_at"), "email": r.get("email"), "api_key": r.get("api_key"), "plan": r.get("plan"), "quota_exports_per_day": r.get("quota_exports_per_day")})
    import sqlite3
    now = now_iso()
    conn = get_conn()
    cur = conn.cursor()
    try:
        api_key = secrets.token_hex(16)
        plan = "basic"
        quota_exports = 5
        today = datetime.datetime.utcnow().date().isoformat()
        cur.execute("INSERT INTO users(username, display_name, created_at, api_key, plan, quota_exports_per_day, exports_used_today, last_quota_reset) VALUES(?, ?, ?, ?, ?, ?, ?, ?)", (body.username, body.display_name, now, api_key, plan, quota_exports, 0, today))
        conn.commit()
        uid = cur.lastrowid
    except sqlite3.IntegrityError:
        conn.close()
        return error_response(400, "VALIDATION_ERROR", "用户名已存在")
    cur.execute("SELECT id, username, display_name, created_at, api_key, plan, quota_exports_per_day FROM users WHERE id = ?", (uid,))
    r = cur.fetchone()
    conn.close()
    return ok({"id": r[0], "username": r[1], "display_name": r[2], "created_at": r[3], "api_key": r[4], "plan": r[5], "quota_exports_per_day": r[6]})

@router.get("/users")
def list_users(page: int = Query(1, ge=1), size: int = Query(20, ge=1, le=100), search: Optional[str] = None):
    q = SB.table("users").select("id,username,display_name,created_at,email", count="exact")
    if search:
        q = q.or_(f"username.ilike.%{search}%,display_name.ilike.%{search}%,email.ilike.%{search}%")
    offset = (page - 1) * size
    res = q.order("id", desc=True).range(offset, offset + size - 1).execute()
    items = getattr(res, "data", None) or []
    total = getattr(res, "count", 0) or len(items)
    return ok({"items": items, "page": page, "size": size, "total": total, "pages": math.ceil(total / size) if size else 0})

@router.get("/users/{user_id}")
def user_detail(user_id: int):
    ures = SB.table("users").select("id,username,display_name,created_at").eq("id", user_id).limit(1).execute()
    u = (getattr(ures, "data", None) or [])
    if not u:
        return error_response(404, "NOT_FOUND", "资源不存在")
    fres = SB.table("user_follows").select("id", count="exact").eq("user_id", user_id).execute()
    follows = getattr(fres, "count", 0) or 0
    r = u[0]
    return ok({"id": r.get("id"), "username": r.get("username"), "display_name": r.get("display_name"), "created_at": r.get("created_at"), "follows": follows})

@router.get("/products/{product_id}/followers")
def product_followers(product_id: int):
    p = get_product(product_id)
    if not p:
        return error_response(404, "NOT_FOUND", "资源不存在")
    links = SB.table("user_follows").select("user_id").eq("product_id", product_id).order("id", desc=True).execute()
    uids = [x.get("user_id") for x in (getattr(links, "data", None) or [])]
    if not uids:
        return ok([])
    users = SB.table("users").select("id,username,display_name,auth_uid").in_("auth_uid", uids).execute()
    items = [{"id": u.get("id"), "username": u.get("username"), "display_name": u.get("display_name") } for u in (getattr(users, "data", None) or [])]
    return ok(items)

@router.get("/pools/public/products")
def list_public_pool_products(page: int = Query(1, ge=1), size: int = Query(20, ge=1, le=100), search: Optional[str] = None, category: Optional[str] = None):
    if SB:
        pres = SB.table("pools").select("id").eq("is_public", True).limit(1).execute()
        pitems = getattr(pres, "data", None) or []
        if not pitems:
            return ok({"items": [], "page": page, "size": size, "total": 0, "pages": 0})
        pool_id = pitems[0]["id"]
        offset = (page - 1) * size
        links_q = SB.table("pool_products").select("product_id", count="exact").eq("pool_id", pool_id).order("id", desc=True)
        links = links_q.range(offset, offset + size - 1).execute()
        ids = [x.get("product_id") for x in (getattr(links, "data", None) or [])]
        total = getattr(links, "count", 0) or 0
        if not ids:
            return ok({"items": [], "page": page, "size": size, "total": total, "pages": math.ceil(total / size) if size else 0})
        pq = SB.table("products").select("*").in_("id", ids)
        if search:
            pq = pq.ilike("name", f"%{search}%")
        if category:
            pq = pq.eq("category", category)
        products = getattr(pq.execute(), "data", None) or []
        items = [{"id": r.get("id"), "name": r.get("name"), "url": r.get("url"), "category": r.get("category"), "last_updated": r.get("updated_at")} for r in products]
        return ok({"items": items, "page": page, "size": size, "total": total, "pages": math.ceil(total / size) if size else 0})
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id FROM pools WHERE is_public = 1 LIMIT 1")
    pool = cur.fetchone()
    offset = (page - 1) * size
    if not pool:
        where = []
        params: List[Any] = []
        if search:
            where.append("name LIKE ?")
            params.append(f"%{search}%")
        if category:
            where.append("category = ?")
            params.append(category)
        where_sql = f"WHERE {' AND '.join(where)}" if where else ""
        total = cur.execute(f"SELECT COUNT(*) FROM products {where_sql}", params).fetchone()[0]
        cur.execute(f"SELECT id,name,url,category,last_updated FROM products {where_sql} ORDER BY id DESC LIMIT ? OFFSET ?", params + [size, offset])
        items = [{"id": r[0], "name": r[1], "url": r[2], "category": r[3], "last_updated": r[4]} for r in cur.fetchall()]
        conn.close()
        return ok({"items": items, "page": page, "size": size, "total": total, "pages": math.ceil(total / size) if size else 0})
    pool_id = pool[0]
    total = cur.execute("SELECT COUNT(*) FROM pool_products WHERE pool_id = ?", (pool_id,)).fetchone()[0]
    cur.execute(
        "SELECT p.id, p.name, p.url, p.category, p.last_updated FROM pool_products pp JOIN products p ON pp.product_id = p.id WHERE pp.pool_id = ? ORDER BY pp.id DESC LIMIT ? OFFSET ?",
        (pool_id, size, offset),
    )
    items = [{"id": r[0], "name": r[1], "url": r[2], "category": r[3], "last_updated": r[4]} for r in cur.fetchall()]
    conn.close()
    return ok({"items": items, "page": page, "size": size, "total": total, "pages": math.ceil(total / size) if size else 0})

@router.get("/pools/public/categories")
def list_public_pool_categories():
    if SB:
        pres = SB.table("pools").select("id").eq("is_public", True).limit(1).execute()
        pitems = getattr(pres, "data", None) or []
        if not pitems:
            return ok([])
        pool_id = pitems[0]["id"]
        links = SB.table("pool_products").select("product_id").eq("pool_id", pool_id).order("id", desc=True).execute()
        ids = [x.get("product_id") for x in (getattr(links, "data", None) or [])]
        if not ids:
            return ok([])
        products = getattr(SB.table("products").select("category").in_("id", ids).execute(), "data", None) or []
        cats = sorted(list({(r.get("category") or "").strip() for r in products if r.get("category")}))
        return ok(cats)
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT category FROM products WHERE category IS NOT NULL AND category <> ''")
    cats = [r[0] for r in cur.fetchall()]
    conn.close()
    return ok(sorted(cats))

@router.post("/pools/public/products")
def add_product_to_public_pool(body: PoolAddProduct):
    if not get_product(body.product_id):
        return error_response(404, "NOT_FOUND", "资源不存在")
    if SB:
        pres = SB.table("pools").select("id").eq("is_public", True).limit(1).execute()
        pitems = getattr(pres, "data", None) or []
        if not pitems:
            cres = SB.table("pools").insert({"name": "public", "is_public": True}).select("id").execute()
            pool_id = (getattr(cres, "data", None) or [{}])[0].get("id")
        else:
            pool_id = pitems[0]["id"]
        try:
            SB.table("pool_products").insert({"pool_id": pool_id, "product_id": body.product_id}).execute()
        except Exception:
            return error_response(400, "VALIDATION_ERROR", "已在公共池")
        return ok({"pool": "public", "product_id": body.product_id})
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id FROM pools WHERE is_public = 1 LIMIT 1")
    pool = cur.fetchone()
    if not pool:
        cur.execute("INSERT INTO pools(name, is_public) VALUES(?, ?)", ("public", 1))
        conn.commit()
        pool_id = cur.lastrowid
    else:
        pool_id = pool[0]
    try:
        cur.execute("INSERT INTO pool_products(pool_id, product_id) VALUES(?, ?)", (pool_id, body.product_id))
        conn.commit()
    except Exception:
        conn.close()
        return error_response(400, "VALIDATION_ERROR", "已在公共池")
    conn.close()
    return ok({"pool": "public", "product_id": body.product_id})

@router.post("/users/{user_id}/select_from_pool")
def user_select_from_pool(user_id: int, body: SelectFromPoolBody):
    if not get_product(body.product_id):
        return error_response(404, "NOT_FOUND", "资源不存在")
    now = now_iso()
    if SB:
        uid = get_auth_uid(user_id)
        if not uid:
            return error_response(404, "NOT_FOUND", "用户不存在")
        try:
            SB.table("user_follows").insert({"user_id": uid, "product_id": body.product_id, "created_at": now}).execute()
        except Exception:
            return error_response(400, "VALIDATION_ERROR", "已选择/关注")
        return ok({"user_id": user_id, "product_id": body.product_id})
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("INSERT INTO user_follows(user_id, product_id, created_at) VALUES(?, ?, ?)", (user_id, body.product_id, now))
        conn.commit()
    except Exception:
        conn.close()
        return error_response(400, "VALIDATION_ERROR", "已选择/关注")
    conn.close()
    return ok({"user_id": user_id, "product_id": body.product_id})

@router.post("/collections")
def create_collection(body: CollectionCreate):
    now = now_iso()
    ures = SB.table("users").select("auth_uid").eq("id", body.owner_user_id).limit(1).execute()
    owner = (getattr(ures, "data", None) or [])
    if not owner:
        return error_response(404, "NOT_FOUND", "用户不存在")
    owner_uid = owner[0].get("auth_uid")
    cres = SB.table("collections").insert({"name": body.name, "owner_user_id": owner_uid, "created_at": now}).select("id").execute()
    cid = (getattr(cres, "data", None) or [{}])[0].get("id")
    SB.table("collection_members").insert({"collection_id": cid, "user_id": owner_uid, "role": "admin"}).execute()
    return ok({"id": cid, "name": body.name, "owner_user_id": body.owner_user_id, "created_at": now})

@router.get("/users/{user_id}/collections")
def list_user_collections(
    user_id: int,
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    search: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    min_members: Optional[int] = None,
    max_members: Optional[int] = None,
    owner_only: Optional[bool] = None,
    owner_id: Optional[int] = None,
    min_products: Optional[int] = None,
    max_products: Optional[int] = None,
    sort_by: Optional[str] = None,
    sort_order: Optional[str] = None,
):
    uid = get_auth_uid(user_id)
    if not uid:
        return ok({"items": [], "page": page, "size": size, "total": 0, "pages": 0})
    links = SB.table("collection_members").select("collection_id").eq("user_id", uid).execute()
    ids = [x.get("collection_id") for x in (getattr(links, "data", None) or [])]
    if not ids:
        return ok({"items": [], "page": page, "size": size, "total": 0, "pages": 0})
    offset = (page - 1) * size
    q = SB.table("collections").select("id,name,created_at,owner_user_id", count="exact").in_("id", ids)
    if search:
        q = q.ilike("name", f"%{search}%")
    # base order by created_at desc unless overridden by sort
    q = q.order("created_at", desc=True)
    res = q.range(offset, offset + size - 1).execute()
    items = getattr(res, "data", None) or []
    total = getattr(res, "count", 0) or len(items)
    # owner filters
    if owner_only:
        items = [r for r in items if str(r.get("owner_user_id") or "") == str(uid)]
    if owner_id is not None:
        owner_uid = get_auth_uid(int(owner_id)) if owner_id else None
        items = [r for r in items if (owner_uid and str(r.get("owner_user_id") or "") == str(owner_uid))]
    # date filters
    if start_date:
        items = [r for r in items if str(r.get("created_at"))[:10] >= start_date]
    if end_date:
        items = [r for r in items if str(r.get("created_at"))[:10] <= end_date]
    cids = [int(r.get("id")) for r in items]
    members_count: Dict[int, int] = {}
    products_count: Dict[int, int] = {}
    last_updated_map: Dict[int, str] = {}
    if cids:
        # member counts
        mres = SB.table("collection_members").select("collection_id").in_("collection_id", cids).execute()
        for r in (getattr(mres, "data", None) or []):
            cid = int(r.get("collection_id"))
            members_count[cid] = members_count.get(cid, 0) + 1
        # product counts and last updated
        pres = SB.table("collection_products").select("collection_id,product_id").in_("collection_id", cids).execute()
        cp = getattr(pres, "data", None) or []
        by_cid: Dict[int, List[int]] = {}
        pid_set: set = set()
        for r in cp:
            cid = int(r.get("collection_id"))
            pid = int(r.get("product_id"))
            by_cid.setdefault(cid, []).append(pid)
            pid_set.add(pid)
        if pid_set:
            prows = getattr(SB.table("products").select("id,updated_at").in_("id", list(pid_set)).execute(), "data", None) or []
            updated_map: Dict[int, str] = {int(p.get("id")): str(p.get("updated_at") or "") for p in prows}
            for cid, pids in by_cid.items():
                products_count[cid] = len(pids)
                # compute max updated_at
                last = ""
                for pid in pids:
                    up = updated_map.get(pid) or ""
                    if up and (not last or up > last):
                        last = up
                last_updated_map[cid] = last
    # members filters
    if min_members is not None:
        items = [r for r in items if members_count.get(int(r.get("id")), 0) >= int(min_members)]
    if max_members is not None:
        items = [r for r in items if members_count.get(int(r.get("id")), 0) <= int(max_members)]
    # products filters
    if min_products is not None:
        items = [r for r in items if products_count.get(int(r.get("id")), 0) >= int(min_products)]
    if max_products is not None:
        items = [r for r in items if products_count.get(int(r.get("id")), 0) <= int(max_products)]
    # sorting
    if sort_by in {"last_updated", "created_at", "name"}:
        reverse = True if sort_order == "desc" else False
        if sort_by == "last_updated":
            items.sort(key=lambda r: (last_updated_map.get(int(r.get("id"))) or ""), reverse=reverse)
        elif sort_by == "created_at":
            items.sort(key=lambda r: str(r.get("created_at") or ""), reverse=reverse)
        else:
            items.sort(key=lambda r: str(r.get("name") or ""), reverse=reverse)
    total = len(items)
    return ok({"items": items, "page": page, "size": size, "total": total, "pages": math.ceil(total / size) if size else 0})
    return ok({"items": items, "page": page, "size": size, "total": total, "pages": math.ceil(total / size) if size else 0})

@router.get("/collections/{collection_id}")
def collection_detail(collection_id: int):
    cres = SB.table("collections").select("id,name,owner_user_id,created_at").eq("id", collection_id).limit(1).execute()
    citems = getattr(cres, "data", None) or []
    if not citems:
        return error_response(404, "NOT_FOUND", "资源不存在")
    c = citems[0]
    links = SB.table("collection_products").select("product_id").eq("collection_id", collection_id).order("id", desc=True).execute()
    pids = [x.get("product_id") for x in (getattr(links, "data", None) or [])]
    products = []
    if pids:
        pres = SB.table("products").select("*").in_("id", pids).execute()
        products = [{"id": r.get("id"), "name": r.get("name"), "url": r.get("url"), "category": r.get("category"), "last_updated": r.get("updated_at")} for r in (getattr(pres, "data", None) or [])]
    ms = SB.table("collection_members").select("user_id,role").eq("collection_id", collection_id).execute()
    uids = [x.get("user_id") for x in (getattr(ms, "data", None) or [])]
    members = []
    if uids:
        ures = SB.table("users").select("id,username,display_name").in_("id", uids).execute()
        udata = getattr(ures, "data", None) or []
        roles = {x.get("user_id"): x.get("role") for x in (getattr(ms, "data", None) or [])}
        members = [{"id": u.get("id"), "username": u.get("username"), "display_name": u.get("display_name"), "role": roles.get(u.get("id"))} for u in udata]
    return ok({"id": c.get("id"), "name": c.get("name"), "owner_user_id": c.get("owner_user_id"), "created_at": c.get("created_at"), "products": products, "members": members})

@router.post("/collections/{collection_id}/products")
def add_collection_product(collection_id: int, body: CollectionAddProduct):
    if not get_product(body.product_id):
        return error_response(404, "NOT_FOUND", "资源不存在")
    cres = SB.table("collections").select("id").eq("id", collection_id).limit(1).execute()
    if not (getattr(cres, "data", None) or []):
        return error_response(404, "NOT_FOUND", "资源不存在")
    try:
        SB.table("collection_products").insert({"collection_id": collection_id, "product_id": body.product_id}).execute()
    except Exception:
        return error_response(400, "VALIDATION_ERROR", "已在集合")
    return ok({"collection_id": collection_id, "product_id": body.product_id})

@router.delete("/collections/{collection_id}/products/{product_id}")
def remove_collection_product(collection_id: int, product_id: int):
    SB.table("collection_products").delete().eq("collection_id", collection_id).eq("product_id", product_id).execute()
    return ok({"collection_id": collection_id, "product_id": product_id})

@router.post("/collections/{collection_id}/share")
def share_collection(collection_id: int, body: CollectionShare):
    cres = SB.table("collections").select("id").eq("id", collection_id).limit(1).execute()
    if not (getattr(cres, "data", None) or []):
        return error_response(404, "NOT_FOUND", "资源不存在")
    ures = SB.table("users").select("auth_uid").eq("id", body.user_id).limit(1).execute()
    if not (getattr(ures, "data", None) or []):
        return error_response(404, "NOT_FOUND", "用户不存在")
    role = body.role if body.role in {"admin", "editor", "viewer"} else "editor"
    try:
        SB.table("collection_members").insert({"collection_id": collection_id, "user_id": (getattr(ures, "data", None) or [{}])[0].get("auth_uid"), "role": role}).execute()
    except Exception:
        return error_response(400, "VALIDATION_ERROR", "成员已存在")
    return ok({"collection_id": collection_id, "user_id": body.user_id, "role": role})

@router.get("/collections/{collection_id}/export.xlsx")
def export_collection_xlsx(collection_id: int, api_key: Optional[str] = None, start_date: Optional[str] = None, end_date: Optional[str] = None):
    cres = SB.table("collections").select("id,name").eq("id", collection_id).limit(1).execute()
    citems = getattr(cres, "data", None) or []
    if not citems:
        return error_response(404, "NOT_FOUND", "资源不存在")
    user = get_user_by_api_key(api_key)
    if user:
        reset_user_quota_if_needed(int(user["id"]))
        quota = user.get("quota_exports_per_day") or 0
        used = user.get("exports_used_today") or 0
        if quota and used >= quota:
            return error_response(429, "QUOTA_EXCEEDED", "导出额度已用尽")
    links = SB.table("collection_products").select("product_id").eq("collection_id", collection_id).execute()
    pids = [x.get("product_id") for x in (getattr(links, "data", None) or [])]
    pres = SB.table("products").select("id,name,url,category").in_("id", pids).execute()
    products = getattr(pres, "data", None) or []
    try:
        from openpyxl import Workbook
    except Exception:
        conn.close()
        return error_response(501, "DEPENDENCY_MISSING", "缺少openpyxl依赖，无法导出Excel")
    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)
    for p in products:
        ws = wb.create_sheet(title=str(p.get("name"))[:31] or f"P{p.get('id')}")
        ws.append(["product_id", "product_name", "url", "category", "price_id", "price", "created_at"])
        q = SB.table("prices").select("id,price,created_at").eq("product_id", p.get("id")).order("created_at", desc=True)
        if start_date:
            q = q.gte("created_at", start_date)
        if end_date:
            q = q.lte("created_at", end_date + " 23:59:59")
        rows = getattr(q.execute(), "data", None) or []
        for r in rows:
            ws.append([p.get("id"), p.get("name"), p.get("url"), p.get("category"), r.get("id"), r.get("price"), r.get("created_at")])
    import io
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"collection_{collection_id}.xlsx"
    if user:
        SB.table("users").update({"exports_used_today": (used or 0) + 1, "last_quota_reset": datetime.datetime.utcnow().date().isoformat()}).eq("id", int(user["id"])).execute()
    return Response(content=bio.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

@router.get("/alerts")
def list_alerts(user_id: Optional[int] = None, product_id: Optional[int] = None):
    q = SB.table("alerts").select("*").order("id", desc=True)
    if user_id is not None:
        uid = get_auth_uid(int(user_id))
        if uid:
            q = q.eq("user_id", uid)
    if product_id is not None:
        q = q.eq("product_id", product_id)
    res = q.execute()
    items = getattr(res, "data", None) or []
    return ok(items)

@router.post("/alerts")
def create_alert(body: AlertCreate):
    now = now_iso()
    if body.rule_type not in {"price_below", "price_above", "percent_drop", "percent_rise"}:
        return error_response(400, "VALIDATION_ERROR", "规则类型无效")
    if not get_product(body.product_id):
        return error_response(404, "NOT_FOUND", "资源不存在")
    uid = get_auth_uid(body.user_id)
    if not uid:
        return error_response(404, "NOT_FOUND", "用户不存在")
    res = SB.table("alerts").insert({"user_id": uid, "product_id": body.product_id, "rule_type": body.rule_type, "threshold": body.threshold, "percent": body.percent, "status": "active", "created_at": now, "updated_at": now, "channel": (body.channel or "inapp"), "cooldown_minutes": (body.cooldown_minutes or 60), "target": body.target}).select("*").execute()
    data = getattr(res, "data", None) or []
    return ok(data[0] if data else {})

@router.post("/alerts/{alert_id}/status")
def update_alert_status(alert_id: int, body: AlertStatusUpdate):
    if body.status not in {"active", "paused"}:
        return error_response(400, "VALIDATION_ERROR", "状态无效")
    now = now_iso()
    SB.table("alerts").update({"status": body.status, "updated_at": now}).eq("id", alert_id).execute()
    res = SB.table("alerts").select("*").eq("id", alert_id).limit(1).execute()
    data = getattr(res, "data", None) or []
    if not data:
        return error_response(404, "NOT_FOUND", "资源不存在")
    return ok(data[0])

@router.delete("/alerts/{alert_id}")
def delete_alert(alert_id: int):
    SB.table("alerts").delete().eq("id", alert_id).execute()
    return ok({"id": alert_id})

@router.get("/alerts/{alert_id}/events")
def list_alert_events(alert_id: int, page: int = Query(1, ge=1), size: int = Query(20, ge=1, le=100), status: Optional[str] = None):
    offset = (page - 1) * size
    q = SB.table("alert_events").select("*", count="exact").eq("alert_id", alert_id).order("id", desc=True)
    if status:
        q = q.eq("status", status)
    res = q.range(offset, offset + size - 1).execute()
    items = getattr(res, "data", None) or []
    total = getattr(res, "count", 0) or len(items)
    return ok({"items": items, "page": page, "size": size, "total": total, "pages": math.ceil(total / size) if size else 0})

@router.post("/alerts/{alert_id}/update")
def update_alert(alert_id: int, threshold: Optional[float] = None, channel: Optional[str] = None, cooldown_minutes: Optional[int] = None):
    now = now_iso()
    payload: Dict[str, Any] = {}
    if threshold is not None:
        payload["threshold"] = threshold
    if channel is not None:
        payload["channel"] = channel
    if cooldown_minutes is not None:
        payload["cooldown_minutes"] = cooldown_minutes
    if not payload:
        return ok({"id": alert_id})
    payload["updated_at"] = now
    SB.table("alerts").update(payload).eq("id", alert_id).execute()
    return ok({"id": alert_id})

@router.post("/alerts/{alert_id}/target")
def update_alert_target(alert_id: int, target: str):
    now = now_iso()
    SB.table("alerts").update({"target": target, "updated_at": now}).eq("id", alert_id).execute()
    return ok({"id": alert_id, "target": target})

@router.get("/users/{user_id}/follows")
def list_user_follows(user_id: int):
    uid = get_auth_uid(user_id)
    if not uid:
        return ok([])
    links = SB.table("user_follows").select("product_id").eq("user_id", uid).order("id", desc=True).execute()
    ids = [x.get("product_id") for x in (getattr(links, "data", None) or [])]
    if not ids:
        return ok([])
    pres = SB.table("products").select("*").in_("id", ids).execute()
    items = [{"id": r.get("id"), "name": r.get("name"), "url": r.get("url"), "category": r.get("category"), "last_updated": r.get("updated_at")} for r in (getattr(pres, "data", None) or [])]
    return ok(items)

@router.get("/users/{user_id}/preferences")
def get_user_preferences(user_id: int):
    res = SB.table("user_preferences").select("*").eq("user_id", user_id).limit(1).execute()
    rows = getattr(res, "data", None) or []
    if not rows:
        return ok({"trend_ma_window": 10, "trend_bb_on": True})
    r = rows[0]
    return ok({"trend_ma_window": r.get("trend_ma_window", 10), "trend_bb_on": bool(r.get("trend_bb_on", True))})

@router.post("/users/{user_id}/preferences")
def update_user_preferences(user_id: int, body: PreferencesUpdate):
    now = now_iso()
    payload: Dict[str, Any] = {"user_id": user_id, "updated_at": now}
    if body.trend_ma_window is not None:
        payload["trend_ma_window"] = int(body.trend_ma_window)
    if body.trend_bb_on is not None:
        payload["trend_bb_on"] = bool(body.trend_bb_on)
    # upsert by user_id
    existing = SB.table("user_preferences").select("id").eq("user_id", user_id).limit(1).execute()
    rows = getattr(existing, "data", None) or []
    if rows:
        SB.table("user_preferences").update(payload).eq("user_id", user_id).execute()
    else:
        payload["created_at"] = now
        SB.table("user_preferences").insert(payload).execute()
    return ok({"user_id": user_id})

@router.post("/users/{user_id}/follows")
def add_follow(user_id: int, body: FollowCreate):
    now = now_iso()
    if not get_product(body.product_id):
        return error_response(404, "NOT_FOUND", "资源不存在")
    try:
        uid = get_auth_uid(user_id)
        if not uid:
            return error_response(404, "NOT_FOUND", "用户不存在")
        SB.table("user_follows").insert({"user_id": uid, "product_id": body.product_id, "created_at": now}).execute()
    except Exception:
        return error_response(400, "VALIDATION_ERROR", "已关注")
    return ok({"user_id": user_id, "product_id": body.product_id})

@router.delete("/users/{user_id}/follows/{product_id}")
def remove_follow(user_id: int, product_id: int):
    uid = get_auth_uid(user_id)
    if not uid:
        return error_response(404, "NOT_FOUND", "用户不存在")
    SB.table("user_follows").delete().eq("user_id", uid).eq("product_id", product_id).execute()
    return ok({"user_id": user_id, "product_id": product_id})

@router.post("/users/{sender_id}/pushes")
def create_push(sender_id: int, body: PushCreate):
    now = now_iso()
    if not get_product(body.product_id):
        return error_response(404, "NOT_FOUND", "资源不存在")
    s_uid = get_auth_uid(sender_id)
    r_uid = get_auth_uid(body.recipient_id)
    if not r_uid:
        return error_response(404, "NOT_FOUND", "接收者不存在")
    res = SB.table("pushes").insert({"sender_id": s_uid, "recipient_id": r_uid, "product_id": body.product_id, "message": body.message, "status": "pending", "created_at": now, "updated_at": now}).select("*").execute()
    data = getattr(res, "data", None) or []
    return ok(data[0] if data else {})

@router.get("/users/{user_id}/pushes")
def list_pushes(user_id: int, box: Optional[str] = None):
    uid = get_auth_uid(user_id)
    if not uid:
        return ok([])
    q = SB.table("pushes").select("*").order("id", desc=True)
    if box == "outbox":
        q = q.eq("sender_id", uid)
    else:
        q = q.eq("recipient_id", uid)
    res = q.execute()
    items = getattr(res, "data", None) or []
    return ok(items)

@router.post("/pushes/{push_id}/status")
def update_push_status(push_id: int, body: PushUpdate):
    now = now_iso()
    if body.status not in {"accepted", "rejected"}:
        return error_response(400, "VALIDATION_ERROR", "状态无效")
    SB.table("pushes").update({"status": body.status, "updated_at": now}).eq("id", push_id).execute()
    res = SB.table("pushes").select("*").eq("id", push_id).limit(1).execute()
    data = getattr(res, "data", None) or []
    if not data:
        return error_response(404, "NOT_FOUND", "资源不存在")
    return ok(data[0])

@router.get("/spider/tasks")
def list_tasks(status: Optional[str] = None, product_id: Optional[int] = None):
    q = SB.table("tasks").select("*").order("priority", desc=True).order("id", desc=True)
    if status:
        q = q.eq("status", status)
    if product_id:
        q = q.eq("product_id", product_id)
    res = q.execute()
    items = getattr(res, "data", None) or []
    return ok(items)

@router.get("/spider/tasks/next")
def next_task():
    q = SB.table("tasks").select("*").eq("status", "pending").order("priority", desc=True).order("scheduled_at", desc=False).order("id", desc=False).limit(1)
    res = q.execute()
    items = getattr(res, "data", None) or []
    return ok(items[0] if items else None)

@router.post("/spider/tasks/next/execute")
def execute_next_task():
    now = now_iso()
    q = SB.table("tasks").select("*").eq("status", "pending").order("priority", desc=True).order("scheduled_at", desc=False).order("id", desc=False).limit(1)
    res = q.execute()
    items = getattr(res, "data", None) or []
    if not items:
        return error_response(404, "NOT_FOUND", "无待执行任务")
    t = items[0]
    tid = int(t.get("id"))
    pid = t.get("product_id")
    if pid is None:
        return error_response(400, "VALIDATION_ERROR", "任务缺少产品ID")
    p = get_product(int(pid))
    if not p:
        return error_response(404, "NOT_FOUND", "资源不存在")
    SB.table("tasks").update({"status": "running", "updated_at": now, "started_at": now}).eq("id", tid).execute()
    rlast = SB.table("prices").select("price,created_at").eq("product_id", pid).order("created_at", desc=True).limit(1).execute()
    last_rows = getattr(rlast, "data", None) or []
    last_price = float(last_rows[0]["price"]) if last_rows else None
    fetched = try_fetch_price(p["url"]) if p and p.get("url") else None
    base = fetched if (isinstance(fetched, float) and fetched > 0) else (last_price if last_price is not None else random.uniform(50.0, 200.0))
    price = round(base * (1 + (0 if fetched else random.uniform(-0.03, 0.03))), 2)
    prev = SB.table("prices").select("price,created_at").eq("product_id", pid).order("created_at", desc=True).limit(1).execute()
    prev_rows = getattr(prev, "data", None) or []
    prev_price = float(prev_rows[0]["price"]) if prev_rows else None
    prev_minute = str(prev_rows[0]["created_at"])[:16] if prev_rows else None
    curr_minute = now[:16]
    if not prev_rows or not (prev_price == price and prev_minute == curr_minute):
        SB.table("prices").insert({"product_id": pid, "price": price, "created_at": now}).execute()
        evaluate_alerts_for_product(int(pid), price, now)
    SB.table("products").update({"updated_at": now}).eq("id", int(pid)).execute()
    SB.table("tasks").update({"status": "completed", "updated_at": now, "completed_at": now}).eq("id", tid).execute()
    t2 = SB.table("tasks").select("*").eq("id", tid).limit(1).execute()
    rows = getattr(t2, "data", None) or []
    r = rows[0] if rows else {}
    return ok({"id": r.get("id"), "product_id": r.get("product_id"), "status": r.get("status"), "created_at": r.get("created_at"), "updated_at": r.get("updated_at")})

@router.post("/spider/tasks")
def create_task(body: TaskCreate, api_key: Optional[str] = Header(None, alias="X-API-Key")):
    now = now_iso()
    if not SB:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("INSERT INTO tasks(product_id, status, priority, created_at, updated_at) VALUES(?, 'pending', ?, ?, ?)", (body.product_id, int(body.priority or 0), now, now))
        conn.commit()
        tid = cur.lastrowid
        conn.close()
        return ok({"id": tid, "product_id": body.product_id, "status": "pending", "created_at": now, "updated_at": now})
    if _is_node_paused():
        return error_response(409, "NODE_PAUSED", "节点暂停，拒绝创建任务")
    created_by = None
    if api_key and isinstance(api_key, str):
        u = get_user_by_api_key(api_key)
        if u:
            today = datetime.datetime.utcnow().date().isoformat()
            ur = SB.table("users").select("quota_tasks_per_day,tasks_created_today,last_tasks_quota_reset").eq("id", int(u.get("id"))).limit(1).execute()
            ud = (getattr(ur, "data", None) or [])
            limit = int((ud[0].get("quota_tasks_per_day") or 20)) if ud else 20
            used = int((ud[0].get("tasks_created_today") or 0)) if ud else 0
            last = (ud[0].get("last_tasks_quota_reset") if ud else None)
            if last != today:
                SB.table("users").update({"tasks_created_today": 0, "last_tasks_quota_reset": today}).eq("id", int(u.get("id"))).execute()
                used = 0
            if used >= limit:
                return error_response(429, "QUOTA_EXCEEDED", "任务创建额度已用尽")
            created_by = int(u.get("id"))
    payload = {"product_id": body.product_id, "status": "pending", "created_at": now, "updated_at": now, "scheduled_at": now, "priority": int(body.priority or 0)}
    if created_by is not None:
        payload["created_by_user_id"] = created_by
    res = SB.table("tasks").insert(payload).select("id").execute()
    data = getattr(res, "data", None) or []
    tid = (data[0] or {}).get("id") if data else None
    if created_by is not None:
        SB.table("users").update({"tasks_created_today": (used if 'used' in locals() else 0) + 1, "last_tasks_quota_reset": datetime.datetime.utcnow().date().isoformat()}).eq("id", created_by).execute()
    return ok({"id": tid, "product_id": body.product_id, "status": "pending", "created_at": now, "updated_at": now})

@router.post("/spider/tasks/{task_id}/execute")
def execute_task(task_id: int):
    now = now_iso()
    if not SB:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT * FROM tasks WHERE id = ?", (task_id,))
        t = cur.fetchone()
        if not t:
            conn.close()
            return error_response(404, "NOT_FOUND", "资源不存在")
        pid = t["product_id"]
        if pid is None:
            conn.close()
            return error_response(400, "VALIDATION_ERROR", "任务缺少产品ID")
        p = get_product(int(pid))
        if not p:
            conn.close()
            return error_response(404, "NOT_FOUND", "资源不存在")
        cur.execute("UPDATE tasks SET status='completed', updated_at=?, completed_at=? WHERE id = ?", (now, now, task_id))
        cur.execute("INSERT INTO prices(product_id, price, created_at) VALUES(?, ?, ?)", (pid, 99.0, now))
        conn.commit()
        cur.execute("SELECT * FROM tasks WHERE id = ?", (task_id,))
        r = cur.fetchone()
        conn.close()
        return ok({"id": r["id"], "product_id": r["product_id"], "status": r["status"], "created_at": r["created_at"], "updated_at": r["updated_at"]})
    if _is_node_paused():
        return error_response(409, "NODE_PAUSED", "节点暂停，拒绝执行")
    tres = SB.table("tasks").select("*").eq("id", task_id).limit(1).execute()
    titems = getattr(tres, "data", None) or []
    if not titems:
        return error_response(404, "NOT_FOUND", "资源不存在")
    t = titems[0]
    pid = t.get("product_id")
    if pid is None:
        return error_response(400, "VALIDATION_ERROR", "任务缺少产品ID")
    p = get_product(int(pid))
    if not p:
        return error_response(404, "NOT_FOUND", "资源不存在")
    SB.table("tasks").update({"status": "running", "updated_at": now, "started_at": now}).eq("id", task_id).execute()
    rlast = SB.table("prices").select("price,created_at").eq("product_id", pid).order("created_at", desc=True).limit(1).execute()
    last_rows = getattr(rlast, "data", None) or []
    last_price = float(last_rows[0]["price"]) if last_rows else None
    fetched = try_fetch_price(p["url"]) if p and p.get("url") else None
    base = fetched if (isinstance(fetched, float) and fetched > 0) else (last_price if last_price is not None else random.uniform(50.0, 200.0))
    price = round(base * (1 + (0 if fetched else random.uniform(-0.03, 0.03))), 2)
    prev = SB.table("prices").select("price,created_at").eq("product_id", pid).order("created_at", desc=True).limit(1).execute()
    prev_rows = getattr(prev, "data", None) or []
    prev_price = float(prev_rows[0]["price"]) if prev_rows else None
    prev_minute = str(prev_rows[0]["created_at"])[:16] if prev_rows else None
    curr_minute = now[:16]
    if not prev_rows or not (prev_price == price and prev_minute == curr_minute):
        SB.table("prices").insert({"product_id": pid, "price": price, "created_at": now}).execute()
        evaluate_alerts_for_product(int(pid), price, now)
    SB.table("products").update({"updated_at": now}).eq("id", int(pid)).execute()
    SB.table("tasks").update({"status": "completed", "updated_at": now, "completed_at": now}).eq("id", task_id).execute()
    t2 = SB.table("tasks").select("*").eq("id", task_id).limit(1).execute()
    rows = getattr(t2, "data", None) or []
    r = rows[0] if rows else {}
    return ok({"id": r.get("id"), "product_id": r.get("product_id"), "status": r.get("status"), "created_at": r.get("created_at"), "updated_at": r.get("updated_at")})

def try_fetch_price(url: str) -> Optional[float]:
    try:
        import urllib.request
        resp = urllib.request.urlopen(url, timeout=5)
        html = resp.read().decode("utf-8", errors="ignore")
        import re
        # simple patterns for currency prices
        patterns = [r"\$\s*(\d+(?:\.\d+)?)", r"¥\s*(\d+(?:\.\d+)?)", r"CNY\s*(\d+(?:\.\d+)?)"]
        for pat in patterns:
            m = re.search(pat, html)
            if m:
                val = float(m.group(1))
                return val
        return None
    except Exception:
        return None

def send_email(to_addr: str, subject: str, body: str):
    import smtplib
    from email.mime.text import MIMEText
    host = os.getenv("SMTP_HOST")
    port = int(os.getenv("SMTP_PORT") or "0")
    user = os.getenv("SMTP_USER")
    password = os.getenv("SMTP_PASS")
    from_addr = os.getenv("SMTP_FROM") or user
    if not host or not port or not user or not password or not from_addr:
        raise RuntimeError("SMTP配置缺失")
    msg = MIMEText(body, _charset="utf-8")
    msg["Subject"] = subject
    msg["From"] = from_addr
    msg["To"] = to_addr
    with smtplib.SMTP(host, port, timeout=5) as s:
        try:
            s.starttls()
        except Exception:
            pass
        s.login(user, password)
        s.sendmail(from_addr, [to_addr], msg.as_string())

def send_webhook(url: str, payload: dict):
    import json
    import urllib.request
    import hmac
    import hashlib
    data = json.dumps(payload).encode("utf-8")
    headers = {"Content-Type": "application/json"}
    ts = now_iso()
    secret = os.getenv("ALERT_WEBHOOK_SECRET")
    if secret:
        sig = hmac.new(secret.encode("utf-8"), data, hashlib.sha256).hexdigest()
        headers["X-Signature"] = sig
        headers["X-Timestamp"] = ts
    req = urllib.request.Request(url, data=data, headers=headers, method="POST")
    with urllib.request.urlopen(req, timeout=5) as resp:
        if resp.status < 200 or resp.status >= 300:
            raise RuntimeError(f"Webhook返回状态码{resp.status}")

def evaluate_alerts_for_product(product_id: int, price: float, now: str):
    res = SB.table("alerts").select("id,user_id,rule_type,threshold,percent,channel,cooldown_minutes,last_triggered_at,target").eq("product_id", product_id).eq("status", "active").execute()
    for a in getattr(res, "data", None) or []:
        rt = a.get("rule_type")
        th = a.get("threshold")
        uid = a.get("user_id")
        trig = False
        if rt == "price_below" and th is not None and price <= float(th):
            trig = True
        if rt == "price_above" and th is not None and price >= float(th):
            trig = True
        if trig:
            channel = a.get("channel") or "inapp"
            cooldown = int(a.get("cooldown_minutes") or 60)
            last_ts = a.get("last_triggered_at")
            allow = True
            if last_ts:
                try:
                    last_dt = datetime.datetime.fromisoformat(str(last_ts).replace("Z", ""))
                    now_dt = datetime.datetime.fromisoformat(now.replace("Z", ""))
                    allow = (now_dt - last_dt).total_seconds() >= cooldown * 60
                except Exception:
                    allow = True
            if not allow:
                continue
            status = "ok"
            err = None
            push_id = None
            try:
                if channel == "inapp":
                    pr = SB.table("pushes").insert({"sender_id": None, "recipient_id": uid, "product_id": product_id, "message": f"价格触发: {price}", "status": "pending", "created_at": now, "updated_at": now}).select("id").execute()
                    push_id = (getattr(pr, "data", None) or [{}])[0].get("id")
                elif channel == "email" and a.get("target"):
                    send_email(str(a.get("target")), "价格触发通知", f"商品{product_id} 当前价格 {price}")
                elif channel == "webhook" and a.get("target"):
                    send_webhook(str(a.get("target")), {"product_id": product_id, "price": price, "time": now})
            except Exception as e:
                status = "failed"
                err = str(e)
            SB.table("alert_events").insert({"alert_id": a.get("id"), "product_id": product_id, "user_id": uid, "price": price, "created_at": now, "message": "触发", "channel": channel, "push_id": push_id, "status": status, "error": err, "attempt": 1}).execute()
            SB.table("alerts").update({"last_triggered_at": now, "updated_at": now}).eq("id", a.get("id")).execute()

@router.post("/alert_events/{event_id}/retry")
def retry_alert_event(event_id: int):
    now = now_iso()
    ev = SB.table("alert_events").select("*").eq("id", event_id).limit(1).execute()
    rows = getattr(ev, "data", None) or []
    if not rows:
        return error_response(404, "NOT_FOUND", "事件不存在")
    e = rows[0]
    aid = e.get("alert_id")
    product_id = e.get("product_id")
    uid = e.get("user_id")
    ar = SB.table("alerts").select("channel,target").eq("id", int(aid)).limit(1).execute()
    arows = getattr(ar, "data", None) or []
    channel = (arows[0] or {}).get("channel") if arows else "inapp"
    target = (arows[0] or {}).get("target") if arows else None
    pricer = SB.table("prices").select("price").eq("product_id", int(product_id)).order("created_at", desc=True).limit(1).execute()
    price_rows = getattr(pricer, "data", None) or []
    price = float((price_rows[0] or {}).get("price") or 0)
    status = "ok"
    err = None
    push_id = None
    try:
        if channel == "inapp":
            pr = SB.table("pushes").insert({"sender_id": 0, "recipient_id": uid, "product_id": product_id, "message": f"价格触发(重试): {price}", "status": "pending", "created_at": now, "updated_at": now}).select("id").execute()
            push_id = (getattr(pr, "data", None) or [{}])[0].get("id")
        elif channel == "email" and target:
            send_email(str(target), "价格触发通知(重试)", f"商品{product_id} 当前价格 {price}")
        elif channel == "webhook" and target:
            send_webhook(str(target), {"product_id": product_id, "price": price, "time": now, "retry": True})
    except Exception as e:
        status = "failed"
        err = str(e)
    SB.table("alert_events").insert({"alert_id": aid, "product_id": product_id, "user_id": uid, "price": price, "created_at": now, "message": "重试", "channel": channel, "push_id": push_id, "status": status, "error": err, "attempt": (int(e.get("attempt") or 1) + 1)}).execute()
    return ok({"event_id": event_id, "status": status})

@router.get("/alerts/{alert_id}/events.csv")
def export_alert_events_csv(alert_id: int, status: Optional[str] = None):
    import csv
    import io
    q = SB.table("alert_events").select("*").eq("alert_id", alert_id).order("id", desc=True)
    if status:
        q = q.eq("status", status)
    res = q.execute()
    items = getattr(res, "data", None) or []
    output = io.StringIO()
    w = csv.writer(output)
    w.writerow(["id", "created_at", "price", "channel", "status", "error", "attempt"]) 
    for e in items:
        w.writerow([e.get("id"), e.get("created_at"), e.get("price"), e.get("channel"), e.get("status"), e.get("error"), e.get("attempt")])
    csv_content = output.getvalue()
    filename = f"alert_{alert_id}_events.csv"
    return Response(content=csv_content, media_type="text/csv", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

@router.post("/spider/listing")
def listing(body: ListingRequest):
    try:
        src_path = os.path.join(BASE_DIR, "src")
        if src_path not in sys.path:
            sys.path.append(src_path)
        from utils.url_util import is_valid_url, get_base_url
    except Exception:
        return error_response(500, "INTERNAL_ERROR", "依赖导入失败")
    if not is_valid_url(body.url):
        return error_response(400, "VALIDATION_ERROR", "URL无效")
    base = get_base_url(body.url)
    if base == "://":
        return error_response(400, "VALIDATION_ERROR", "URL无效")
    limit = max(1, min(body.max_items, 50))
    items = [{"title": f"Item {i+1}", "url": body.url, "source": base} for i in range(limit)]
    return ok({"count": len(items), "items": items})

@router.get("/export")
def export_products(product_ids: str, api_key: Optional[str] = None, start_date: Optional[str] = None, end_date: Optional[str] = None):
    ids = [int(x) for x in product_ids.split(",") if x.strip().isdigit()]
    if not ids:
        return error_response(400, "VALIDATION_ERROR", "缺少有效的product_ids")
    user = get_user_by_api_key(api_key)
    if user:
        reset_user_quota_if_needed(int(user["id"]))
        quota = user.get("quota_exports_per_day") or 0
        used = user.get("exports_used_today") or 0
        if quota and used >= quota:
            return error_response(429, "QUOTA_EXCEEDED", "导出额度已用尽")
    import csv
    import io
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["product_id", "product_name", "url", "category", "price_id", "price", "created_at"])
    for pid in ids:
        p = get_product(pid)
        if not p:
            continue
        q = SB.table("prices").select("id,price,created_at").eq("product_id", pid).order("created_at", desc=True)
        if start_date:
            q = q.gte("created_at", start_date)
        if end_date:
            q = q.lte("created_at", end_date + " 23:59:59")
        rows = getattr(q.execute(), "data", None) or []
        for r in rows:
            writer.writerow([pid, p["name"], p["url"], p["category"], r.get("id"), r.get("price"), r.get("created_at")])
    csv_content = output.getvalue()
    filename = "products_export.csv"
    if user:
        SB.table("users").update({"exports_used_today": (used or 0) + 1, "last_quota_reset": datetime.datetime.utcnow().date().isoformat()}).eq("id", int(user["id"])).execute()
    return Response(content=csv_content, media_type="text/csv", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

@router.get("/export.xlsx")
def export_products_xlsx(product_ids: str, api_key: Optional[str] = None):
    ids = [int(x) for x in product_ids.split(",") if x.strip().isdigit()]
    if not ids:
        return error_response(400, "VALIDATION_ERROR", "缺少有效的product_ids")
    try:
        from openpyxl import Workbook
    except Exception:
        return error_response(500, "INTERNAL_ERROR", "缺少 openpyxl 依赖")
    user = get_user_by_api_key(api_key)
    if user:
        reset_user_quota_if_needed(user["id"])
        quota = user["quota_exports_per_day"] or 0
        used = user["exports_used_today"] or 0
        if quota and used >= quota:
            return error_response(429, "QUOTA_EXCEEDED", "导出额度已用尽")
    wb = Workbook()
    # remove default sheet
    default = wb.active
    wb.remove(default)
    header = ["product_id", "product_name", "url", "category", "price_id", "price", "created_at"]
    conn = get_conn()
    cur = conn.cursor()
    for pid in ids:
        p = get_product(pid)
        if not p:
            continue
        ws = wb.create_sheet(title=f"product_{pid}")
        ws.append(header)
        cur.execute("SELECT id, price, created_at FROM prices WHERE product_id = ? ORDER BY created_at DESC", (pid,))
        for r in cur.fetchall():
            ws.append([pid, p["name"], p["url"], p["category"], r["id"], r["price"], r["created_at"]])
    conn.close()
    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    data = bio.getvalue()
    filename = "products_export.xlsx"
    if user:
        conn2 = get_conn()
        cur2 = conn2.cursor()
        cur2.execute("UPDATE users SET exports_used_today = COALESCE(exports_used_today, 0) + 1 WHERE id = ?", (user["id"],))
        conn2.commit()
        conn2.close()
    return Response(content=data, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

@router.post("/graphql")
def graphql_endpoint(payload: dict = Body(...)):
    query: str = payload.get("query") or ""
    variables: dict = payload.get("variables") or {}
    def resp(data):
        return {"data": data}
    if "products" in query and "mutation" not in query:
        page = int(variables.get("page", 1))
        size = int(variables.get("size", 20))
        if SB is not None:
            offset = (page - 1) * size
            sel = SB.table("products").select("*", count="exact").order("id", desc=True).range(offset, offset + size - 1).execute()
            items = [{"id": r.get("id"), "name": r.get("name"), "url": r.get("url"), "category": r.get("category"), "last_updated": r.get("updated_at")} for r in (getattr(sel, "data", None) or [])]
            total = getattr(sel, "count", 0) or len(items)
            return resp({"products": {"items": items, "total": total, "page": page, "size": size}})
        else:
            conn = get_conn()
            cur = conn.cursor()
            total = cur.execute("SELECT COUNT(*) FROM products").fetchone()[0]
            offset = (page - 1) * size
            cur.execute("SELECT * FROM products ORDER BY id DESC LIMIT ? OFFSET ?", (size, offset))
            items = [row_to_product(r) for r in cur.fetchall()]
            conn.close()
            return resp({"products": {"items": items, "total": total, "page": page, "size": size}})
    if "product(" in query and "mutation" not in query:
        pid = int(variables.get("id"))
        p = get_product(pid)
        if not p:
            return resp({"product": None})
        conn = get_conn()
        stats = conn.execute("SELECT COUNT(*) as cnt, MAX(price) as max_price, MIN(price) as min_price, AVG(price) as avg_price FROM prices WHERE product_id = ?", (pid,)).fetchone()
        conn.close()
        p["stats"] = {"count": stats[0], "max_price": stats[1], "min_price": stats[2], "avg_price": stats[3]}
        return resp({"product": p})
    if "productPrices" in query:
        pid = int(variables.get("product_id"))
        if SB is not None:
            res = SB.table("prices").select("*").eq("product_id", pid).order("created_at", desc=True).execute()
            items = [{"id": r.get("id"), "product_id": r.get("product_id"), "price": r.get("price"), "created_at": r.get("created_at")} for r in (getattr(res, "data", None) or [])]
            return resp({"productPrices": items})
        else:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute("SELECT * FROM prices WHERE product_id = ? ORDER BY created_at DESC", (pid,))
            items = [row_to_price(r) for r in cur.fetchall()]
            conn.close()
            return resp({"productPrices": items})
    if "createProduct" in query:
        input = variables.get("input") or {}
        pid = create_product(input.get("name"), input.get("url"), input.get("category"))
        return resp({"createProduct": get_product(pid)})
    if "updateProduct" in query:
        pid = int(variables.get("id"))
        input = variables.get("input") or {}
        now = now_iso()
        if SB is not None:
            SB.table("products").update({"name": input.get("name"), "url": input.get("url"), "category": input.get("category"), "updated_at": now}).eq("id", pid).execute()
            return resp({"updateProduct": get_product(pid)})
        else:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute("UPDATE products SET name = COALESCE(?, name), url = COALESCE(?, url), category = COALESCE(?, category), last_updated = ? WHERE id = ?", (input.get("name"), input.get("url"), input.get("category"), now, pid))
            conn.commit()
            cur.execute("SELECT * FROM products WHERE id = ?", (pid,))
            r = cur.fetchone()
            conn.close()
            return resp({"updateProduct": row_to_product(r) if r else None})
    if "deleteProduct" in query:
        pid = int(variables.get("id"))
        if SB is not None:
            SB.table("prices").delete().eq("product_id", pid).execute()
            SB.table("tasks").delete().eq("product_id", pid).execute()
            SB.table("products").delete().eq("id", pid).execute()
            return resp({"deleteProduct": {"id": pid}})
        else:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute("DELETE FROM prices WHERE product_id = ?", (pid,))
            cur.execute("DELETE FROM tasks WHERE product_id = ?", (pid,))
            cur.execute("DELETE FROM products WHERE id = ?", (pid,))
            conn.commit()
            conn.close()
            return resp({"deleteProduct": {"id": pid}})
    if "getManyProducts" in query:
        ids = variables.get("ids") or []
        items = []
        for pid in ids:
            p = get_product(int(pid))
            if p:
                items.append(p)
        return resp({"getManyProducts": {"items": items}})
    if "alerts" in query and "mutation" not in query:
        user_id = variables.get("user_id")
        product_id = variables.get("product_id")
        q = SB.table("alerts").select("*").order("id", desc=True)
        if user_id is not None:
            q = q.eq("user_id", user_id)
        if product_id is not None:
            q = q.eq("product_id", product_id)
        res = q.execute()
        items = getattr(res, "data", None) or []
        return resp({"alerts": items})
    if "createAlert" in query:
        input = variables.get("input") or {}
        now = now_iso()
        rt = str(input.get("rule_type")) if input.get("rule_type") is not None else ""
        if rt not in {"price_below", "price_above", "percent_drop", "percent_rise"}:
            return resp({"createAlert": None})
        if not get_product(int(input.get("product_id"))):
            return resp({"createAlert": None})
        res = SB.table("alerts").insert({
            "user_id": input.get("user_id"),
            "product_id": input.get("product_id"),
            "rule_type": input.get("rule_type"),
            "threshold": input.get("threshold"),
            "percent": input.get("percent"),
            "status": "active",
            "created_at": now,
            "updated_at": now,
            "channel": input.get("channel") or "inapp",
            "cooldown_minutes": input.get("cooldown_minutes") or 60,
            "target": input.get("target"),
        }).select("*").execute()
        data = getattr(res, "data", None) or []
        return resp({"createAlert": (data[0] if data else None)})
    if "updateAlertStatus" in query:
        alert_id = int(variables.get("id"))
        status = str(variables.get("status"))
        if status not in {"active", "paused"}:
            return resp({"updateAlertStatus": None})
        now = now_iso()
        SB.table("alerts").update({"status": status, "updated_at": now}).eq("id", alert_id).execute()
        res = SB.table("alerts").select("*").eq("id", alert_id).limit(1).execute()
        data = getattr(res, "data", None) or []
        return resp({"updateAlertStatus": (data[0] if data else None)})
    if "deleteAlert" in query:
        alert_id = int(variables.get("id"))
        SB.table("alerts").delete().eq("id", alert_id).execute()
        return resp({"deleteAlert": {"id": alert_id}})
    if "alertEvents" in query and "mutation" not in query:
        alert_id = int(variables.get("alert_id"))
        page = int(variables.get("page", 1))
        size = int(variables.get("size", 20))
        status = variables.get("status")
        offset = (page - 1) * size
        q = SB.table("alert_events").select("*", count="exact").eq("alert_id", alert_id).order("id", desc=True)
        if status:
            q = q.eq("status", status)
        res = q.range(offset, offset + size - 1).execute()
        items = getattr(res, "data", None) or []
        total = getattr(res, "count", 0) or len(items)
        return resp({"alertEvents": {"items": items, "total": total, "page": page, "size": size}})
    if "updateAlert" in query:
        alert_id = int(variables.get("id"))
        input = variables.get("input") or {}
        now = now_iso()
        payload = {}
        if input.get("threshold") is not None:
            payload["threshold"] = input.get("threshold")
        if input.get("channel") is not None:
            payload["channel"] = input.get("channel")
        if input.get("cooldown_minutes") is not None:
            payload["cooldown_minutes"] = input.get("cooldown_minutes")
        if not payload:
            return resp({"updateAlert": {"id": alert_id}})
        payload["updated_at"] = now
        SB.table("alerts").update(payload).eq("id", alert_id).execute()
        return resp({"updateAlert": {"id": alert_id}})
    if "updateAlertTarget" in query:
        alert_id = int(variables.get("id"))
        target = variables.get("target")
        now = now_iso()
        SB.table("alerts").update({"target": target, "updated_at": now}).eq("id", alert_id).execute()
        return resp({"updateAlertTarget": {"id": alert_id, "target": target}})
    if "userCollections" in query and "mutation" not in query:
        user_id = int(variables.get("user_id"))
        page = int(variables.get("page", 1))
        size = int(variables.get("size", 20))
        search = variables.get("search")
        links = SB.table("collection_members").select("collection_id").eq("user_id", user_id).execute()
        ids = [x.get("collection_id") for x in (getattr(links, "data", None) or [])]
        if not ids:
            return resp({"userCollections": {"items": [], "page": page, "size": size, "total": 0}})
        offset = (page - 1) * size
        q = SB.table("collections").select("id,name,created_at", count="exact").in_("id", ids).order("id", desc=True)
        if search:
            q = q.ilike("name", f"%{search}%")
        res = q.range(offset, offset + size - 1).execute()
        items = getattr(res, "data", None) or []
        total = getattr(res, "count", 0) or len(items)
        return resp({"userCollections": {"items": items, "page": page, "size": size, "total": total}})
    if "collection(" in query and "mutation" not in query:
        cid = int(variables.get("id"))
        cres = SB.table("collections").select("id,name,owner_user_id,created_at").eq("id", cid).limit(1).execute()
        citems = getattr(cres, "data", None) or []
        if not citems:
            return resp({"collection": None})
        c = citems[0]
        links = SB.table("collection_products").select("product_id").eq("collection_id", cid).order("id", desc=True).execute()
        pids = [x.get("product_id") for x in (getattr(links, "data", None) or [])]
        products = []
        if pids:
            pres = SB.table("products").select("*").in_("id", pids).execute()
            products = [{"id": r.get("id"), "name": r.get("name"), "url": r.get("url"), "category": r.get("category"), "last_updated": r.get("updated_at")} for r in (getattr(pres, "data", None) or [])]
        ms = SB.table("collection_members").select("user_id,role").eq("collection_id", cid).execute()
        uids = [x.get("user_id") for x in (getattr(ms, "data", None) or [])]
        members = []
        if uids:
            ures = SB.table("users").select("id,username,display_name").in_("id", uids).execute()
            udata = getattr(ures, "data", None) or []
            roles = {x.get("user_id"): x.get("role") for x in (getattr(ms, "data", None) or [])}
            members = [{"id": u.get("id"), "username": u.get("username"), "display_name": u.get("display_name"), "role": roles.get(u.get("id"))} for u in udata]
        return resp({"collection": {"id": c.get("id"), "name": c.get("name"), "owner_user_id": c.get("owner_user_id"), "created_at": c.get("created_at"), "products": products, "members": members}})
    if "addCollectionProduct" in query:
        cid = int(variables.get("collection_id"))
        pid = int(variables.get("product_id"))
        if not get_product(pid):
            return resp({"addCollectionProduct": None})
        cres = SB.table("collections").select("id").eq("id", cid).limit(1).execute()
        if not (getattr(cres, "data", None) or []):
            return resp({"addCollectionProduct": None})
        try:
            SB.table("collection_products").insert({"collection_id": cid, "product_id": pid}).execute()
        except Exception:
            return resp({"addCollectionProduct": None})
        return resp({"addCollectionProduct": {"collection_id": cid, "product_id": pid}})
    if "removeCollectionProduct" in query:
        cid = int(variables.get("collection_id"))
        pid = int(variables.get("product_id"))
        SB.table("collection_products").delete().eq("collection_id", cid).eq("product_id", pid).execute()
        return resp({"removeCollectionProduct": {"collection_id": cid, "product_id": pid}})
    if "shareCollection" in query:
        cid = int(variables.get("collection_id"))
        input = variables.get("input") or {}
        uid = int(input.get("user_id"))
        role = input.get("role") if input.get("role") in {"admin", "editor", "viewer"} else "editor"
        cres = SB.table("collections").select("id").eq("id", cid).limit(1).execute()
        if not (getattr(cres, "data", None) or []):
            return resp({"shareCollection": None})
        ures = SB.table("users").select("id").eq("id", uid).limit(1).execute()
        if not (getattr(ures, "data", None) or []):
            return resp({"shareCollection": None})
        try:
            SB.table("collection_members").insert({"collection_id": cid, "user_id": uid, "role": role}).execute()
        except Exception:
            return resp({"shareCollection": None})
        return resp({"shareCollection": {"collection_id": cid, "user_id": uid, "role": role}})
    if "userPushes" in query and "mutation" not in query:
        user_id = int(variables.get("user_id"))
        box = variables.get("box")
        q = SB.table("pushes").select("*").order("id", desc=True)
        if box == "outbox":
            q = q.eq("sender_id", user_id)
        else:
            q = q.eq("recipient_id", user_id)
        res = q.execute()
        items = getattr(res, "data", None) or []
        return resp({"userPushes": items})
    if "updatePushStatus" in query:
        push_id = int(variables.get("id"))
        status = str(variables.get("status"))
        if status not in {"accepted", "rejected"}:
            return resp({"updatePushStatus": None})
        now = now_iso()
        SB.table("pushes").update({"status": status, "updated_at": now}).eq("id", push_id).execute()
        res = SB.table("pushes").select("*").eq("id", push_id).limit(1).execute()
        data = getattr(res, "data", None) or []
        return resp({"updatePushStatus": (data[0] if data else None)})
    if "createPush" in query:
        sender_id = int(variables.get("sender_id"))
        input = variables.get("input") or {}
        now = now_iso()
        if not get_product(int(input.get("product_id"))):
            return resp({"createPush": None})
        res = SB.table("pushes").insert({
            "sender_id": sender_id,
            "recipient_id": input.get("recipient_id"),
            "product_id": input.get("product_id"),
            "message": input.get("message"),
            "status": "pending",
            "created_at": now,
            "updated_at": now,
        }).select("*").execute()
        data = getattr(res, "data", None) or []
        return resp({"createPush": (data[0] if data else None)})
    return resp({})

app.include_router(router)
try:
    src_path = os.path.join(BASE_DIR, "src")
    if src_path not in sys.path:
        sys.path.append(src_path)
    from ai.api import router as ai_router  # noqa
    app.include_router(ai_router)
except Exception:
    pass

def main():
    print("Hello from spider!")

if __name__ == "__main__":
    main()
@router.get("/products/{product_id}/export/xlsx")
def export_product_prices_xlsx(product_id: int, start_date: Optional[str] = None, end_date: Optional[str] = None, api_key: Optional[str] = Header(None)):
    p = get_product(product_id)
    if not p:
        return error_response(404, "NOT_FOUND", "资源不存在")
    user = get_user_by_api_key(api_key)
    if user:
        reset_user_quota_if_needed(int(user["id"]))
        quota = user.get("quota_exports_per_day") or 0
        used = user.get("exports_used_today") or 0
        if quota and (used + 1) > quota:
            return error_response(429, "QUOTA_EXCEEDED", "导出额度已用尽")
    from openpyxl import Workbook
    import io
    wb = Workbook()
    ws = wb.active
    ws.title = (p.get("name") or f"product_{product_id}")[:31].replace("/", "-")
    ws.append(["product_id", "product_name", "url", "category", "price_id", "price", "created_at"])
    q = SB.table("prices").select("id,price,created_at").eq("product_id", product_id).order("created_at", desc=True)
    if start_date:
        q = q.gte("created_at", start_date)
    if end_date:
        q = q.lte("created_at", end_date + " 23:59:59")
    rows = getattr(q.execute(), "data", None) or []
    for r in rows:
        ws.append([product_id, p["name"], p["url"], p["category"], r.get("id"), r.get("price"), r.get("created_at")])
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    if user:
        SB.table("users").update({"exports_used_today": (user.get("exports_used_today") or 0) + 1, "last_quota_reset": datetime.datetime.utcnow().date().isoformat()}).eq("id", int(user["id"])).execute()
    filename = f"product_{product_id}_prices.xlsx"
    return Response(content=content, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})
